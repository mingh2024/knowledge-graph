"""
Build js/knowledge-graph-data.json from the Excel source table.

Source of truth: data/knowledge-graph-source.xlsx
  - 节点表   (Node_ID, 名称, 类型, 所属领域, 二级分类, 定义, 标签, ...)
  - 关系表   (Source_ID, Source_Name, Source_Type, Target_ID, Target_Name, Target_Type, 关系类型, 强度, 说明)
  - 课程映射 (Course_ID, 课程名称, 课程类别, 建议阶段, 推荐排序, 前置课程, 覆盖节点数, 覆盖重点主题, 覆盖节点ID, 课程说明, 时间节点)
  - 应用映射 (应用ID, 应用场景, 关联节点数, 关键关联节点, 关联课程, 应用说明)

Run after editing the Excel file to regenerate js/knowledge-graph-data.json:
    python3 build_data_from_excel.py
"""
import json
import os
import re
import sys

try:
    import openpyxl
except ImportError:
    sys.exit("Missing dependency: run 'pip install openpyxl' first")

BASE = os.path.dirname(os.path.abspath(__file__))
SOURCE_XLSX = os.path.join(BASE, "data", "knowledge-graph-source.xlsx")
OUT_JSON = os.path.join(BASE, "js", "knowledge-graph-data.json")

YEAR_NAMES = ["大一", "大二", "大三", "大四"]
YEAR_COLORS = {"大一": "#5B9BD5", "大二": "#4ECDC4", "大三": "#9B59B6", "大四": "#F39C12"}

SPLIT_RE = re.compile(r"[、,，;；\s]+")


def split_ids(cell):
    if not cell:
        return []
    return [x for x in SPLIT_RE.split(str(cell).strip()) if x]


def last_year_token(cell):
    """'大三或大四' -> '大四'; '大一' -> '大一'."""
    if not cell:
        return ""
    parts = re.split(r"或|/|、", str(cell).strip())
    return parts[-1].strip() if parts else str(cell).strip()


def cell_str(v):
    return "" if v is None else str(v).strip()


def load_workbook():
    if not os.path.exists(SOURCE_XLSX):
        sys.exit(f"Source Excel not found: {SOURCE_XLSX}")
    return openpyxl.load_workbook(SOURCE_XLSX, data_only=True)


def rows(ws, header_row=1):
    headers = [cell_str(ws.cell(row=header_row, column=c).value) for c in range(1, ws.max_column + 1)]
    for r in range(header_row + 1, ws.max_row + 1):
        values = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
        if all(v is None or cell_str(v) == "" for v in values):
            continue
        yield dict(zip(headers, values))


def main():
    wb = load_workbook()
    warnings = []

    # ---- 节点表 ----
    nodes = {}
    for row in rows(wb["节点表"]):
        nid = cell_str(row.get("Node_ID"))
        if not nid:
            continue
        nodes[nid] = {
            "id": nid,
            "name": cell_str(row.get("名称")),
            "type": cell_str(row.get("类型")),
            "year": "",
            "course": "",
            "definition": cell_str(row.get("定义")),
            "domain": cell_str(row.get("所属领域")),
            "subCategory": cell_str(row.get("二级分类")),
            "tags": cell_str(row.get("标签")),
            "relatedNodeIds": [],
            "appKeyNodes": [],
            "appCourses": [],
            "appDescription": "",
        }

    # ---- 关系表 ----
    relations = []
    related = {}
    for row in rows(wb["关系表"]):
        sid = cell_str(row.get("Source_ID"))
        tid = cell_str(row.get("Target_ID"))
        if not sid or not tid:
            continue
        for missing in (sid, tid):
            if missing not in nodes:
                warnings.append(f"关系表引用了节点表中不存在的 ID：{missing}（已自动补齐占位节点，请在 Excel 中修正）")
                nodes[missing] = {
                    "id": missing,
                    "name": f"未定义节点_{missing}",
                    "type": "待补充",
                    "year": "",
                    "course": "",
                    "definition": "关系表中已引用，但节点表未提供定义；为保证图数据库可导入而自动补齐。",
                    "domain": "待补充",
                    "subCategory": "待补充",
                    "tags": "待补充 | 自动补齐",
                    "relatedNodeIds": [],
                    "appKeyNodes": [],
                    "appCourses": [],
                    "appDescription": "",
                }
        strength = row.get("强度")
        try:
            strength = float(strength) if strength is not None else 1
            if strength == int(strength):
                strength = int(strength)
        except (TypeError, ValueError):
            strength = 1
        relations.append({
            "sourceId": sid,
            "targetId": tid,
            "type": cell_str(row.get("关系类型")),
            "strength": strength,
            "description": cell_str(row.get("说明")),
        })
        related.setdefault(sid, set()).add(tid)
        related.setdefault(tid, set()).add(sid)

    for nid, n in nodes.items():
        n["relatedNodeIds"] = sorted(related.get(nid, set()))

    # ---- 课程映射 ----
    courses = {}  # Course_ID -> info
    for row in rows(wb["课程映射"]):
        cid = cell_str(row.get("Course_ID"))
        if not cid:
            continue
        try:
            order = float(row.get("推荐排序"))
        except (TypeError, ValueError):
            order = 9999
        courses[cid] = {
            "id": cid,
            "name": cell_str(row.get("课程名称")),
            "order": order,
            "year": last_year_token(row.get("时间节点")),
            "nodeIds": split_ids(row.get("覆盖节点ID")),
        }

    # id -> list of course ids covering it, for picking the "primary" course
    coverage = {}
    for cid, c in courses.items():
        for nid in c["nodeIds"]:
            coverage.setdefault(nid, []).append(cid)

    # ---- 应用映射 ----
    app_info = {}
    for row in rows(wb["应用映射"]):
        aid = cell_str(row.get("应用ID"))
        if not aid:
            continue
        app_info[aid] = {
            "keyNodes": split_ids(row.get("关键关联节点")),
            "courseNames": split_ids(row.get("关联课程")),
            "description": cell_str(row.get("应用说明")),
        }
    # ---- assign year/course per node ----
    unassigned = []
    for nid, n in nodes.items():
        if n["type"] == "课程":
            c = courses.get(nid)
            if c:
                n["year"] = c["year"]
            # A course can itself be grouped under a broader course (e.g. a
            # capstone like 毕业论文 that lists other courses in its own
            # 覆盖节点ID) - that's its `course`, distinct from its own year.
            cids = coverage.get(nid, [])
            n["course"] = courses[cids[0]]["name"] if cids else ""
        else:
            if n["type"] == "应用":
                info = app_info.get(nid)
                if info:
                    n["appKeyNodes"] = info["keyNodes"]
                    n["appCourses"] = info["courseNames"]
                    n["appDescription"] = info["description"] or n["definition"]
            cids = coverage.get(nid, [])
            if cids:
                # Primary course = the first one that lists this node, in
                # 课程映射 sheet order (not by 推荐排序 - a node can be
                # covered by multiple courses, and the sheet's row order is
                # what determines which one "owns" it).
                best = courses[cids[0]]
                n["year"] = best["year"]
                n["course"] = best["name"]
            else:
                unassigned.append(nid)

    # Fallback: a node with no direct course coverage inherits the majority
    # year among its directly-related nodes (iterated, since neighbors may
    # themselves need to resolve first). This only fills in `year`, never
    # `course` — the real fix is adding the node to a course's 覆盖节点ID.
    if unassigned:
        remaining = set(unassigned)
        for _ in range(10):
            if not remaining:
                break
            progressed = False
            for nid in list(remaining):
                neighbor_years = [nodes[r]["year"] for r in related.get(nid, ()) if nodes.get(r, {}).get("year")]
                if not neighbor_years:
                    continue
                counts = {}
                for y in neighbor_years:
                    counts[y] = counts.get(y, 0) + 1
                best_year = max(YEAR_NAMES, key=lambda y: (counts.get(y, 0), -YEAR_NAMES.index(y)))
                if counts.get(best_year, 0) == 0:
                    continue
                nodes[nid]["year"] = best_year
                remaining.discard(nid)
                progressed = True
            if not progressed:
                break
        if remaining:
            for nid in remaining:
                nodes[nid]["year"] = YEAR_NAMES[0]
            warnings.append(
                f"{len(remaining)} 个节点既未被任何课程覆盖、也没有已知年级的关联节点，"
                f"已默认归入「{YEAR_NAMES[0]}」：{', '.join(sorted(remaining))}"
            )
        warnings.append(
            f"{len(unassigned)} 个节点未被任何课程覆盖（课程映射-覆盖节点ID 中找不到），"
            f"其年级是按关联节点推断出来的，建议在 Excel 课程映射表里补上覆盖关系：{', '.join(unassigned[:20])}"
            + (" ..." if len(unassigned) > 20 else "")
        )

    # ---- years[] ----
    years_out = []
    for yname in YEAR_NAMES:
        year_courses = sorted(
            (c for c in courses.values() if c["year"] == yname),
            key=lambda c: c["order"],
        )
        courses_out = []
        year_node_ids = []
        for c in year_courses:
            cnode_ids = sorted(nid for nid, n in nodes.items() if n["course"] == c["name"])
            courses_out.append({"id": c["id"], "name": c["name"], "nodeIds": cnode_ids})
        year_node_ids = sorted(nid for nid, n in nodes.items() if n["year"] == yname)
        years_out.append({
            "name": yname,
            "color": YEAR_COLORS[yname],
            "courses": courses_out,
            "nodeIds": year_node_ids,
        })

    data = {
        "nodes": sorted(nodes.values(), key=lambda n: n["id"]),
        "relations": relations,
        "years": years_out,
        "meta": {"nodeCount": len(nodes), "relationCount": len(relations)},
    }

    os.makedirs(os.path.dirname(OUT_JSON), exist_ok=True)
    with open(OUT_JSON, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

    print(f"Wrote {OUT_JSON}")
    print(f"  nodes: {len(nodes)}  relations: {len(relations)}")
    if warnings:
        print(f"\n{len(warnings)} warning(s):")
        for w in warnings:
            print(f"  - {w}")
        sys.exit(1 if any("补齐占位节点" in w for w in warnings) else 0)


if __name__ == "__main__":
    main()
